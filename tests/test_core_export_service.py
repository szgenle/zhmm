"""Tests for :mod:`zhmm.core.export_service`."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from zhmm.core.errors import StorageError, ValidationError
from zhmm.core.export_service import CN_HEADS, ExportService
from zhmm.core.models import PasswordEntry


@pytest.fixture
def entries() -> list[PasswordEntry]:
    return [
        PasswordEntry(
            id=1,
            role="工作",
            userID="alice",
            pwd="p1",
            phone="13800138000",
            email="a@x.com",
            url="https://example.com",
            desc="line1\r\nline2",
            utime=1000,
        ),
        PasswordEntry(id=2, role="个人", userID="bob", pwd="中文密码", phone="", email="", url="", desc="", utime=2000),
    ]


class TestExportImportRoundtrip:
    def test_roundtrip(self, tmp_path: Path, entries):
        f = tmp_path / "out.xlsx"
        ExportService.export_xlsx(f, entries)
        loaded = ExportService.import_xlsx(f)
        # 字段完全一致（含换行转义还原）
        assert [e.to_dict() for e in loaded] == [e.to_dict() for e in entries]

    def test_newlines_preserved(self, tmp_path: Path):
        f = tmp_path / "nl.xlsx"
        entry = PasswordEntry(id=1, userID="u", pwd="p", desc="a\r\nb\nc", utime=1)
        ExportService.export_xlsx(f, [entry])
        loaded = ExportService.import_xlsx(f)
        assert loaded[0].desc == "a\r\nb\nc"

    def test_empty_entries(self, tmp_path: Path):
        f = tmp_path / "empty.xlsx"
        ExportService.export_xlsx(f, [])
        loaded = ExportService.import_xlsx(f)
        assert loaded == []


class TestImportErrors:
    def test_missing_file(self, tmp_path: Path):
        with pytest.raises(StorageError):
            ExportService.import_xlsx(tmp_path / "nope.xlsx")

    def test_bad_header_raises(self, tmp_path: Path):
        f = tmp_path / "bad.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["wrong", "headers"])
        wb.save(str(f))
        wb.close()
        with pytest.raises(ValidationError):
            ExportService.import_xlsx(f)


class TestImportQuirks:
    def test_phone_float_suffix_stripped(self, tmp_path: Path):
        f = tmp_path / "phone.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(list(CN_HEADS))
        # 手机被 openpyxl 读成 float 的典型场景：这里先模拟 str "13800138000.0"
        ws.append([1, "个人", "u", "p", "13800138000.0", "", "", "", 0])
        wb.save(str(f))
        wb.close()
        loaded = ExportService.import_xlsx(f)
        assert loaded[0].phone == "13800138000"

    def test_skip_completely_empty_row(self, tmp_path: Path):
        f = tmp_path / "blank.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(list(CN_HEADS))
        ws.append([1, "个人", "u", "p", "", "", "", "", 0])
        ws.append([None, None, None, None, None, None, None, None, None])
        wb.save(str(f))
        wb.close()
        loaded = ExportService.import_xlsx(f)
        assert len(loaded) == 1

    def test_invalid_int_becomes_zero(self, tmp_path: Path):
        f = tmp_path / "bad_int.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(list(CN_HEADS))
        ws.append(["not-a-number", "个人", "u", "p", "", "", "", "", "xyz"])
        wb.save(str(f))
        wb.close()
        loaded = ExportService.import_xlsx(f)
        assert loaded[0].id == 0
        assert loaded[0].utime == 0


class TestTotpExport:
    def test_totp_secret_is_stripped_on_export(self, tmp_path: Path):
        """导出 xlsx 故意不包含 totp_secret，仅保留算法标识。"""
        f = tmp_path / "with_totp.xlsx"
        entry = PasswordEntry(
            id=1,
            userID="u",
            pwd="p",
            totp_secret="JBSWY3DPEHPK3PXP",
            totp_algo="SM3",
            totp_digits=8,
            totp_period=60,
            utime=1,
        )
        ExportService.export_xlsx(f, [entry])
        loaded = ExportService.import_xlsx(f)
        # secret 已抹除
        assert loaded[0].totp_secret == ""
        # 算法标识保留
        assert loaded[0].totp_algo == "SM3"
        assert loaded[0].totp_digits == 8
        assert loaded[0].totp_period == 60

    def test_legacy_xlsx_without_totp_columns_still_imports(self, tmp_path: Path):
        """老 xlsx 只有 9 列核心表头时应正常导入，TOTP 字段回默认。"""
        f = tmp_path / "legacy.xlsx"
        wb = Workbook()
        ws = wb.active
        # 仅写入原 9 列表头
        legacy_heads = ["ID", "类别", "账号", "密码", "手机", "邮箱", "网站", "备注", "更新时间"]
        ws.append(legacy_heads)
        ws.append([1, "个人", "u", "p", "", "", "", "", 1])
        wb.save(str(f))
        wb.close()
        loaded = ExportService.import_xlsx(f)
        assert len(loaded) == 1
        # TOTP 字段回默认
        assert loaded[0].totp_secret == ""
        assert loaded[0].totp_algo == ""


class TestTagsExport:
    def test_tags_roundtrip(self, tmp_path: Path):
        """含 tags 的条目导出后导入应完全一致。"""
        f = tmp_path / "with_tags.xlsx"
        entry = PasswordEntry(id=1, userID="u", pwd="p", utime=1, tags=["work", "prod"])
        ExportService.export_xlsx(f, [entry])
        loaded = ExportService.import_xlsx(f)
        assert loaded[0].tags == ["work", "prod"]

    def test_empty_tags_export_as_blank(self, tmp_path: Path):
        """空 tags 导出为空串，导入回空列表。"""
        f = tmp_path / "blank_tags.xlsx"
        entry = PasswordEntry(id=1, userID="u", pwd="p", utime=1, tags=[])
        ExportService.export_xlsx(f, [entry])
        loaded = ExportService.import_xlsx(f)
        assert loaded[0].tags == []

    def test_legacy_xlsx_without_tags_column_still_imports(self, tmp_path: Path):
        """旧 xlsx 无「标签」列时应正常导入，tags 回空列表。"""
        f = tmp_path / "legacy_no_tags.xlsx"
        wb = Workbook()
        ws = wb.active
        legacy_heads = ["ID", "类别", "账号", "密码", "手机", "邮箱", "网站", "备注", "更新时间"]
        ws.append(legacy_heads)
        ws.append([1, "个人", "u", "p", "", "", "", "", 1])
        wb.save(str(f))
        wb.close()
        loaded = ExportService.import_xlsx(f)
        assert loaded[0].tags == []

    def test_import_tags_semicolon_split_and_normalize(self, tmp_path: Path):
        """手写 xlsx 分号单元格能被拆分，空段被过滤。"""
        f = tmp_path / "manual_tags.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(list(CN_HEADS))
        row = [1, "个人", "u", "p", "", "", "", "", 1, "", 6, 30, "a;b; ;a; c"]
        ws.append(row)
        wb.save(str(f))
        wb.close()
        loaded = ExportService.import_xlsx(f)
        # 去空 / 去重 / strip
        assert loaded[0].tags == ["a", "b", "c"]
