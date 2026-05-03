"""Excel 导入导出服务。

从老 `data_exporter.py` 迁移到 core 层，统一：
- 输入输出是 `PasswordEntry` 列表，不再是 TypedDict dict
- 失败抛 `StorageError` / `ValidationError`，不再 print + return None
- 不再有 `print` 侧效，由调用方决定如何汇报

特殊字符转义规则与历史一致：`\r` ↔ `[r]`，`\n` ↔ `[n]`，便于跨 sheet 编辑。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from zhmm.core.errors import StorageError, ValidationError
from zhmm.core.models import PasswordEntry, normalize_tags

CN_HEADS: tuple[str, ...] = (
    "ID",
    "类别",
    "账号",
    "密码",
    "手机",
    "邮箱",
    "网站",
    "备注",
    "更新时间",
    # 下列 TOTP 配置仅导出算法标识，故意不导出原始 secret
    "TOTP算法",
    "TOTP位数",
    "TOTP周期",
    # 标签：多个标签以英文分号 `;` 分隔，空标签会被过滤
    "标签",
)
EN_HEADS: tuple[str, ...] = (
    "id",
    "role",
    "userID",
    "pwd",
    "phone",
    "email",
    "url",
    "desc",
    "utime",
    "totp_algo",
    "totp_digits",
    "totp_period",
    "tags",
)
# 核心列：导入时仅对这 9 列做必存校验，以兼容旧 xlsx（TOTP / 标签 都是可选列）。
_CORE_CN_HEADS: tuple[str, ...] = CN_HEADS[:9]
_INT_FIELDS = {"id", "utime", "totp_digits", "totp_period"}
# 标签在 Excel 单元格内的分隔符；与 UI Chip 编辑器和 normalize_tags 约定一致。
_TAG_SEP: str = ";"
# 注意：``PasswordEntry.history``（密码历史版本）刻意不列入 CN_HEADS / EN_HEADS。
# Excel 通道不导出（避免明文历史密码落盘扩散），也不导入（导入后 history 为空）。


def _escape(value: str) -> str:
    return value.replace("\r", "[r]").replace("\n", "[n]")


def _unescape(value: str) -> str:
    return value.replace("[r]", "\r").replace("[n]", "\n")


class ExportService:
    """Excel 导入导出。静态方法形式。"""

    # ---------- 导出 ----------

    @staticmethod
    def export_xlsx(file_path: str | Path, entries: list[PasswordEntry]) -> None:
        """把条目导出为 xlsx。

        Raises:
            StorageError: 文件写入失败
        """
        wb = Workbook()
        ws = wb.active
        if ws is None:  # pragma: no cover - openpyxl 保证
            raise StorageError("failed to create worksheet")
        ws.title = "密码数据"
        ws.append(list(CN_HEADS))
        for e in entries:
            row: list[str] = []
            d = e.to_dict()
            for key in EN_HEADS:
                v = d.get(key, "")
                if key == "tags":
                    # tags 是 list，用分号分隔为字符串；空列表 → ""
                    if isinstance(v, (list | tuple)):
                        row.append(_escape(_TAG_SEP.join(str(t) for t in v)))
                    else:
                        row.append(_escape(str(v) if v is not None else ""))
                    continue
                row.append(_escape(str(v) if v is not None else ""))
            ws.append(row)
        try:
            wb.save(str(file_path))
        except OSError as ex:
            raise StorageError(f"write failed: {file_path}: {ex}") from ex
        finally:
            wb.close()

    # ---------- 导入 ----------

    @staticmethod
    def import_xlsx(file_path: str | Path) -> list[PasswordEntry]:
        """从 xlsx 导入条目。

        Raises:
            StorageError: 文件不存在 / 读取失败
            ValidationError: 表头不匹配
        """
        p = Path(file_path)
        if not p.exists():
            raise StorageError(f"file not found: {p}")
        try:
            wb = load_workbook(str(p), data_only=True)
        except OSError as ex:
            raise StorageError(f"read failed: {p}: {ex}") from ex

        try:
            ws = wb.active
            if ws is None:
                raise ValidationError("worksheet is empty")
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                raise ValidationError("header row is missing")
            headers = [str(h) if h is not None else "" for h in header_row]
            missing = [h for h in _CORE_CN_HEADS if h not in headers]
            if missing:
                raise ValidationError(f"missing columns: {missing}")
            index_map = {h: i for i, h in enumerate(headers)}

            entries: list[PasswordEntry] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None or all(v is None or v == "" for v in row):
                    continue
                data: dict[str, object] = {}
                for cn, en in zip(CN_HEADS, EN_HEADS, strict=True):
                    col = index_map.get(cn, -1)
                    value = row[col] if 0 <= col < len(row) else None
                    if value is None:
                        if en == "tags":
                            data[en] = []
                        else:
                            data[en] = 0 if en in _INT_FIELDS else ""
                        continue
                    text = str(value)
                    # openpyxl 对纯数字列常会返回 float，手机号需去掉小数尾
                    if en == "phone" and text.endswith(".0"):
                        text = text[:-2]
                    text = _unescape(text)
                    if en == "tags":
                        # 标签：分号分隔文本 → list[str]（统一过 normalize 跟 PasswordEntry 对齐）
                        data[en] = normalize_tags(text)
                    elif en in _INT_FIELDS:
                        try:
                            data[en] = int(float(text)) if text else 0
                        except (ValueError, TypeError):
                            data[en] = 0
                    else:
                        data[en] = text
                entries.append(PasswordEntry.from_dict(data))
            return entries
        finally:
            wb.close()
