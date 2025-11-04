from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from zhmm.data.sm_data_types import ZhmmDict


class DataImporter:
    @staticmethod
    def import_from_file(xlsx_file_path: str) -> list[ZhmmDict] | None:
        """
        从文件导入数据
        检查文件格式
        检查文件内容
        导入数据
        返回导入结果
        """
        cn_heads = ["ID", "类别", "账号", "密码", "手机", "邮箱", "网站", "备注", "更新时间"]
        en_heads = [
            "id",
            "role",
            "userID",
            "pwd",
            "phone",
            "email",
            "url",
            "desc",
            "utime",
        ]

        try:
            # 使用 openpyxl 读取 Excel
            wb = load_workbook(xlsx_file_path, data_only=True)
            ws = wb.active

            if ws is None:
                print("无法读取工作表")
                return None

            # 获取表头（第一行）
            headers = [cell.value for cell in ws[1]]

            # 检查列名是否匹配
            if not all(col in headers for col in cn_heads):
                print("Excel文件列名不匹配")
                return None

            # 创建列索引映射
            col_index_map = {header: idx for idx, header in enumerate(headers)}

            data: list[ZhmmDict] = []
            # 从第二行开始读取数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = {}
                for i, cn_col in enumerate(cn_heads):
                    col_idx = col_index_map.get(cn_col, -1)
                    if col_idx == -1:
                        continue

                    cell_value = row[col_idx] if col_idx < len(row) else None

                    # 处理空值和转换为字符串
                    if cell_value is None:
                        value = ""
                    else:
                        value = str(cell_value)

                    # 特殊处理手机号列（去除浮点数后缀）
                    if cn_col == "手机" and value.endswith(".0"):
                        value = value[:-2]

                    # 还原特殊字符
                    value = value.replace("[r]", "\r").replace("[n]", "\n")

                    # 转换ID和更新时间为整数
                    if en_heads[i] in ["id", "utime"]:
                        try:
                            item[en_heads[i]] = int(float(value)) if value else None  # type: ignore
                        except (ValueError, TypeError):
                            item[en_heads[i]] = None  # type: ignore
                    else:
                        item[en_heads[i]] = value
                data.append(item)  # type: ignore

            wb.close()
            print(f"成功从 {xlsx_file_path} 导入 {len(data)} 条数据")
            return data

        except FileNotFoundError:
            print(f"文件不存在: {xlsx_file_path}")
            return None
        except ImportError as e:
            print(f"缺少必要的库，请安装 openpyxl: {str(e)}")
            return None
        except Exception as e:
            import traceback
            print(f"导入Excel文件失败: {str(e)}")
            print(traceback.format_exc())
            return None


class DataExporter:
    """数据导出工具类"""

    @staticmethod
    def export_xlsx(file_path, data):
        """导出xlsx文件"""
        cn_heads = ["ID", "类别", "账号", "密码", "手机", "邮箱", "网站", "备注", "更新时间"]
        en_heads = [
            "id",
            "role",
            "userID",
            "pwd",
            "phone",
            "email",
            "url",
            "desc",
            "utime",
        ]

        try:
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active

            if ws is None:
                print("无法创建工作表")
                return False

            ws.title = "密码数据"

            # 写入表头
            ws.append(cn_heads)

            # 写入数据行
            for item in data:
                row_data = []
                for key in en_heads:
                    value = str(item[key]) if key in item else ""
                    # 转义特殊字符
                    value = value.replace("\r", "[r]").replace("\n", "[n]")
                    row_data.append(value)
                ws.append(row_data)

            # 保存文件
            wb.save(file_path)
            wb.close()
            print(f"数据已成功导出到: {file_path}")
            return True
        except Exception as e:
            print(f"导出Excel文件失败: {str(e)}")
            # 降级到 CSV 导出
            try:
                import csv
                csv_path = file_path.replace(".xlsx", ".csv")
                with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(cn_heads)
                    for item in data:
                        row_data = []
                        for key in en_heads:
                            value = str(item[key]) if key in item else ""
                            value = value.replace("\r", "[r]").replace("\n", "[n]")
                            row_data.append(value)
                        writer.writerow(row_data)
                print(f"已改为CSV格式导出到: {csv_path}")
                return True
            except Exception as csv_e:
                print(f"CSV导出也失败: {str(csv_e)}")
            return False
